"""
Scan JPG images from data folder and create Excel sheet with metadata.
"""

import os
from pathlib import Path
from datetime import datetime
from PIL import Image
from PIL.ExifTags import TAGS
import openpyxl
from openpyxl import Workbook

from water_meter_detector import detect_water_meter_type


def get_image_datetime(image_path):
    """
    Extract the date and time when the image was taken from EXIF metadata.
    
    Args:
        image_path: Path to the image file
        
    Returns:
        Formatted datetime string or None if no EXIF data found
    """
    try:
        image = Image.open(image_path)
        exif_data = image._getexif()
        
        if exif_data is not None:
            for tag_id, value in exif_data.items():
                tag = TAGS.get(tag_id, tag_id)
                
                # Look for DateTime or DateTimeOriginal
                if tag in ['DateTime', 'DateTimeOriginal', 'DateTimeDigitized']:
                    return value
        
        return None
    except Exception as e:
        print(f"Error reading EXIF data from {image_path}: {e}")
        return None


def scan_images_to_excel():
    """
    Scan JPG images from data folder and create Excel sheet with metadata.
    """
    # Define paths
    data_folder = Path('data')
    results_folder = Path('results')
    
    # Create results folder if it doesn't exist
    results_folder.mkdir(exist_ok=True)
    
    # Get all JPG files from data folder (case-insensitive)
    jpg_files = []
    seen_files = set()
    for pattern in ['*.jpg', '*.JPG', '*.jpeg', '*.JPEG']:
        for file_path in data_folder.glob(pattern):
            if file_path.name.lower() not in seen_files:
                jpg_files.append(file_path)
                seen_files.add(file_path.name.lower())
    
    if not jpg_files:
        print("No JPG files found in data folder.")
        return
    
    # Sort files by name
    jpg_files.sort()
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Metadata"
    
    # Add headers
    ws['A1'] = 'File Name'
    ws['B1'] = 'Date and Time Taken'
    ws['C1'] = 'Meter Type'
    
    # Make headers bold
    ws['A1'].font = openpyxl.styles.Font(bold=True)
    ws['B1'].font = openpyxl.styles.Font(bold=True)
    ws['C1'].font = openpyxl.styles.Font(bold=True)
    
    # Process each image
    row = 2
    for image_path in jpg_files:
        filename = image_path.name
        datetime_taken = get_image_datetime(image_path)
        meter_type = detect_water_meter_type(image_path)
        
        ws[f'A{row}'] = filename
        ws[f'B{row}'] = datetime_taken if datetime_taken else 'No EXIF data'
        ws[f'C{row}'] = meter_type
        
        print(f"Processed: {filename} - {datetime_taken} - {meter_type}")
        row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    
    # Create output filename with current date
    current_date = datetime.now().strftime('%Y.%m.%dT%H.%M')
    output_filename = f"{current_date}__images.xlsx"
    output_path = results_folder / output_filename
    
    # Save Excel file
    wb.save(output_path)
    print(f"\nExcel file created: {output_path}")
    print(f"Total images processed: {len(jpg_files)}")


if __name__ == "__main__":
    scan_images_to_excel()
